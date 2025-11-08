# -*- coding: utf-8 -*-
"""
核心功能模块：第一步 原始数据分析（与 GUI / 文件输出解耦）
================================================================
本模块只做“原始数据处理”，输入一个 Excel 文件路径，输出计算结果变量，供 GUI
层（例如 main_app/gui_step1_rawdata.py）写入到 Excel/绘图/展示。

满足两点修改：
1) 仅输入路径 -> 返回数据；不负责命名与输出 Excel。
2) 读取 Excel **所有单元格**中的数字（跨所有工作表），而非仅第一列。

主要对外函数：
- process_raw_excel(file_path: str, target_prob: float = 0.97, parse_string_numbers: bool = True)
    -> Step1RawDataResult

返回结果中包含：
- 分类统计表（value, count, proportion）
- 覆盖≥target_prob的最短区间 (min, median, max, interval_prob)
- 计数信息（numbers_count, non_numbers_count, blank_cells）

注意：不依赖任何 GUI；不进行文件写入或绘图。
"""
from __future__ import annotations


from dataclasses import dataclass
from typing import List, Tuple, Optional

from openpyxl import load_workbook

__all__ = [
    "CategoryStatsEntry",
    "Step1RawDataResult",
    "process_raw_excel",
    "read_all_numbers_from_excel",
    "calculate_counts_and_proportions",
    "find_shortest_interval_covering_prob",
]


# ============================
# 数据结构
# ============================
@dataclass
class CategoryStatsEntry:
    value: float
    count: int
    proportion: float


@dataclass
class Step1RawDataResult:
    # 源文件路径（便于上层记录/显示）
    file_path: str
    # 分类统计：按 value 升序
    category_stats: List[CategoryStatsEntry]
    # 覆盖区间摘要
    min_val: Optional[float]
    median_val: Optional[float]
    max_val: Optional[float]
    interval_prob: float
    # 计数信息
    numbers_count: int
    non_numbers_count: int
    blank_cells: int

    @property
    def has_data(self) -> bool:
        return len(self.category_stats) > 0


# ============================
# 读取：遍历所有工作表与所有单元格
# ============================

def _try_parse_float(val) -> Tuple[bool, Optional[float]]:
    """尽力将值解析为浮点数。
    - 原生 int/float -> 接受
    - 字符串 -> 去空白，去千分位逗号，尝试 float
    - 其它类型（bool/datetime/None 等）-> 失败
    返回 (ok, number)
    """
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        # openpyxl 把布尔视为 bool, 避免当作数字
        return True, float(val)
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return False, None
        # 去掉常见千分位逗号
        s = s.replace(",", "")
        try:
            return True, float(s)
        except Exception:
            return False, None
    return False, None


def read_all_numbers_from_excel(
    file_path: str,
    parse_string_numbers: bool = True,
) -> Tuple[List[float], int, int, int]:
    """读取工作簿中**所有工作表**的**所有单元格**里的数字。

    参数：
    - file_path: Excel 路径
    - parse_string_numbers: 是否将可解析的数字字符串也当作数字

    返回：
    - numbers: List[float]        # 收集到的数值
    - numbers_count: int          # 数字单元格数量（含被解析的数字字符串）
    - non_numbers_count: int      # 非空但无法解析为数字的单元格数量
    - blank_cells: int            # 空白单元格数量
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)

    numbers: List[float] = []
    numbers_count = 0
    non_numbers_count = 0
    blank_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    blank_cells += 1
                    continue
                if parse_string_numbers:
                    ok, num = _try_parse_float(v)
                else:
                    ok = isinstance(v, (int, float)) and not isinstance(v, bool)
                    num = float(v) if ok else None

                if ok and num is not None:
                    numbers.append(num)
                    numbers_count += 1
                else:
                    non_numbers_count += 1

    return numbers, numbers_count, non_numbers_count, blank_cells


# ============================
# 统计与区间
# ============================

def calculate_counts_and_proportions(numbers: List[float]) -> List[CategoryStatsEntry]:
    """统计 value -> count 和 proportion（按 value 升序）。"""
    total = len(numbers)
    if total == 0:
        return []
    # 计数
    counts = {}
    for v in numbers:
        counts[v] = counts.get(v, 0) + 1
    # 组装并排序
    entries = [CategoryStatsEntry(value=v, count=c, proportion=c / total) for v, c in counts.items()]
    entries.sort(key=lambda e: e.value)
    return entries


def find_shortest_interval_covering_prob(
    stats: List[CategoryStatsEntry], target_prob: float = 0.97
) -> Tuple[Optional[float], Optional[float], Optional[float], float]:
    """在离散分布上寻找覆盖≥target_prob的最短区间。
    返回 (min, median, max, interval_prob)。若无数据，返回 (None, None, None, 0.0)。
    """
    n = len(stats)
    if n == 0:
        return None, None, None, 0.0

    values = [s.value for s in stats]
    props = [s.proportion for s in stats]

    best_i, best_j = 0, n - 1
    best_span = None

    for i in range(n):
        cum = 0.0
        for j in range(i, n):
            cum += props[j]
            if cum + 1e-12 >= target_prob:
                span = values[j] - values[i]
                if best_span is None or span < best_span:
                    best_span = span
                    best_i, best_j = i, j
                break

    min_val = values[best_i]
    max_val = values[best_j]
    interval_prob = sum(props[best_i: best_j + 1])

    # 区间内的“加权中位”点（按 proportion 权重累积到区间总概率的一半）
    cum = 0.0
    half = 0.5 * interval_prob
    median_val = min_val
    for k in range(best_i, best_j + 1):
        cum += props[k]
        if cum >= half:
            median_val = values[k]
            break

    return min_val, median_val, max_val, interval_prob


# ============================
# 对外主入口
# ============================

def process_raw_excel(
    file_path: str,
    target_prob: float = 0.97,
    parse_string_numbers: bool = True,
) -> Step1RawDataResult:
    """读取 Excel(全表全单元格)，计算分类统计与最短覆盖区间，返回纯数据结果。

    仅做计算，不写任何文件/不画图；供 GUI 层把结果写入 Excel 或展示。
    """
    numbers, n_num, n_non, n_blank = read_all_numbers_from_excel(
        file_path=file_path,
        parse_string_numbers=parse_string_numbers,
    )

    stats = calculate_counts_and_proportions(numbers)
    min_val, median_val, max_val, interval_prob = find_shortest_interval_covering_prob(stats, target_prob)

    return Step1RawDataResult(
        file_path=file_path,
        category_stats=stats,
        min_val=min_val,
        median_val=median_val,
        max_val=max_val,
        interval_prob=interval_prob,
        numbers_count=n_num,
        non_numbers_count=n_non,
        blank_cells=n_blank,
    )
