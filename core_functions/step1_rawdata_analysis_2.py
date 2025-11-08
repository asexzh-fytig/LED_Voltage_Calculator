# -*- coding: utf-8 -*-
"""
模块：core_functions/step1_rawdata_analysis_2.py
功能：读取表格并返回纯“数值单元格”数据（多表、全单元格）。
支持输入格式：
  - .xlsx / .xlsm / .xltx / .xltm （使用 openpyxl，仅读单元格为数值类型的内容）
  - .xls （使用 xlrd，仅读取数值类型单元格）
  - .csv （使用内置 csv 模块；仅当整格为纯数字文本时才转换为数值，含单位/百分号等不转换）

注意：
  - 只读取“数值格式”，不会从非数值文本里提取数字（如"3.2V"、"5%"将被忽略）。
  - 若抽取到的数值总量 > 1,000,000，采用水库抽样（Reservoir Sampling）等概率保留 1,000,000 个；
    否则原样返回（保留逐行结构）。
返回：
    data_numeric : list[list[float]]
        - 若总量 <= 1,000,000：保持“逐行”分组
        - 若总量  > 1,000,000：返回 1,000,000 个样本，每行一个数 [[x], [y], ...]
"""
from __future__ import annotations

import os
import math
import re
import csv
import random
from typing import Iterable, List, Tuple

# -------------------- 格式检测 --------------------
_XLSX_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_XLS_EXTS = {".xls"}
_CSV_EXTS = {".csv"}


# -------------------- 纯数字字符串判断（用于 CSV） --------------------
# 允许：可选空白、可选正负号、十进制（含小数点）或科学计数，整体必须是数字；不允许千分位逗号、百分号、单位等。
_NUMERIC_FULL_RE = re.compile(
    r"""^\s*[+\-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+\-]?\d+)?\s*$"""
)


def _is_numeric_string(s: str) -> bool:
    if s is None:
        return False
    return bool(_NUMERIC_FULL_RE.match(str(s)))


# -------------------- 逐行迭代器（不同格式各自实现） --------------------
def _iter_numeric_rows_xlsx(file_path: str) -> Iterable[List[float]]:
    """openpyxl：遍历所有工作表、所有行，仅保留数值类型单元格。"""
    from openpyxl import load_workbook  # 延迟导入，避免未安装时报错

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cur: List[float] = []
                for cell in row:
                    if isinstance(cell, (int, float)):
                        v = float(cell)
                        if math.isfinite(v):
                            cur.append(v)
                if cur:
                    yield cur
    finally:
        wb.close()


def _iter_numeric_rows_xls(file_path: str) -> Iterable[List[float]]:
    """xlrd：仅支持 .xls；仅读取数值类型单元格。"""
    try:
        import xlrd  # xlrd>=2.0 仅支持 .xls；请确保环境中已安装
    except Exception as e:
        raise ImportError("读取 .xls 需要安装 'xlrd' 库（仅支持旧版 .xls）。") from e

    book = xlrd.open_workbook(file_path, on_demand=True)
    try:
        for sh in book.sheets():
            for r in range(sh.nrows):
                cur: List[float] = []
                row = sh.row(r)
                for c in row:
                    # xlrd.XL_CELL_NUMBER == 2
                    if c.ctype == xlrd.XL_CELL_NUMBER:
                        v = float(c.value)
                        if math.isfinite(v):
                            cur.append(v)
                if cur:
                    yield cur
    finally:
        book.release_resources()


def _iter_numeric_rows_csv(file_path: str) -> Iterable[List[float]]:
    """csv：仅当整格为纯数字文本时才转换为浮点数，其余忽略。"""
    with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            cur: List[float] = []
            for field in row:
                if _is_numeric_string(field):
                    try:
                        v = float(field)
                    except Exception:
                        continue
                    if math.isfinite(v):
                        cur.append(v)
            if cur:
                yield cur


# -------------------- 聚合 + 水库抽样 --------------------
def _collect_with_sampling(
    row_iter: Iterable[List[float]],
    max_points: int,
) -> List[List[float]]:
    """
    将行级迭代器聚合为二维数组；若总量超限，转为水库抽样，返回 [[x], ...]。
    """
    data_numeric: List[List[float]] = []
    reservoir: List[float] = []
    k = int(max_points)
    total = 0
    using_reservoir = False

    for row in row_iter:
        if not row:
            continue

        if using_reservoir:
            for val in row:
                total += 1
                j = random.randint(1, total)
                if j <= k:
                    reservoir[j - 1] = val
        else:
            data_numeric.append(row)
            total += len(row)
            if total > k:
                # 切换到水库模式：将已有所有数 flatten 后随机抽 k 个作为初始水库
                flat_all: List[float] = [v for r in data_numeric for v in r]
                reservoir = random.sample(flat_all, k)
                data_numeric = []  # 释放原结构
                using_reservoir = True

    if using_reservoir:
        return [[x] for x in reservoir]
    else:
        return data_numeric


# -------------------- 对外主函数 --------------------
def read_excel_numeric(file_path: str, max_points: int = 1_000_000) -> List[List[float]]:
    """
    读取文件并返回二维数值数组。
    仅读取“数值格式”：
        - Excel：只接收数值类型单元格（文本单元格一律忽略）
        - CSV：仅当整格为纯数字文本时才转换为浮点数，其余忽略
    超过 max_points 条则水库抽样等概率保留 max_points 条。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext in _XLSX_EXTS:
        rows = _iter_numeric_rows_xlsx(file_path)
    elif ext in _XLS_EXTS:
        rows = _iter_numeric_rows_xls(file_path)
    elif ext in _CSV_EXTS:
        rows = _iter_numeric_rows_csv(file_path)
    else:
        raise ValueError(f"不支持的文件扩展名：{ext}（支持 .xlsx/.xlsm/.xltx/.xltm/.xls/.csv）")

    return _collect_with_sampling(rows, max_points)


# -------------------- 自测 --------------------
if __name__ == "__main__":
    # CSV 字段纯数字 -> 接受；包含单位/百分号 -> 忽略
    print(_is_numeric_string("  -3.25  "))        # True
    print(_is_numeric_string("6.02e-3"))          # True
    print(_is_numeric_string("1,234.5"))          # False（含千分位逗号）
    print(_is_numeric_string("3.2V"))             # False（含单位）
    print(_is_numeric_string("5%"))               # False（含百分号）
