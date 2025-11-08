# -*- coding: utf-8 -*-
"""
第一步：原始数据分析（PyQt版本）
---------------------------------------------
基于 core_functions/step1_rawdata_analysis.py 的纯计算结果，
本界面负责：
- 选择至多5个 Excel 文件（可留空）
- 调用 process_raw_excel 读取全表全单元格数字并计算统计量
- 将结果写入统一汇总 Excel（data_files/output/step1_rawdata_analysis.xlsx）
- 生成每个 LED 的概率分布图 PNG（data_files/output/LEDx概率分布区间图.png）
- 在界面中预览 5 张图；双击某张图可在系统中打开对应 PNG

提示：本文件仅包含 GUI 与文件输出逻辑；核心计算已在 core_functions 中。
"""
import os
import sys
import subprocess
import csv
from pathlib import Path

# PyQt相关导入
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QLineEdit, QPushButton, QProgressBar, QTextEdit, QMessageBox,
                             QFileDialog, QFrame, QScrollArea, QGridLayout)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal
from PyQt5.QtGui import QPixmap, QFont

# 路径
from path_manager import get_step1_dir, get_resources_dir

# 图像/Excel/绘图相关
from openpyxl import Workbook
import matplotlib

matplotlib.use("Agg")  # 使用无界面后端生成 PNG
import matplotlib.pyplot as plt

# 核心计算函数
from core_functions.step1_rawdata_analysis_2 import read_excel_numeric
from core_functions.step1_rawdata_analysis import (
    process_raw_excel,
    Step1RawDataResult,
    CategoryStatsEntry,
)

# 避免中文乱码与负号问题
matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Arial", "DejaVu Sans"]
matplotlib.rcParams["axes.unicode_minus"] = False

# ---------------------------------------------
# 路径工具：定位 data_files/output
# ---------------------------------------------
_DEF_SEARCH_LEVELS = 6


def _find_project_root(start_dir: str) -> str:
    cur = os.path.abspath(start_dir)
    for _ in range(_DEF_SEARCH_LEVELS):
        has_data = os.path.isdir(os.path.join(cur, "data_files"))
        has_main = os.path.isdir(os.path.join(cur, "main_app"))
        has_core = os.path.isdir(os.path.join(cur, "core_functions"))
        if has_data or (has_main and has_core):
            return cur
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent
    return os.path.dirname(os.path.abspath(start_dir))


def _ensure_output_dir() -> str:
    """确保并返回 data_files/step1_rawdata_analysis 目录"""
    return get_step1_dir()


# ---------- 初始化原始CSV文件（每个只包含一个数字0） ----------
def _init_led_csv_files(out_dir: str) -> None:
    """
    软件开始运行时，先创建 5 个只包含一个数字 0 的 CSV 文件：
    LED1_raw_data.csv ~ LED5_raw_data.csv
    """
    os.makedirs(out_dir, exist_ok=True)
    for i in range(1, 6):
        csv_path = os.path.join(out_dir, f"LED{i}_raw_data.csv")
        try:
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([0])
        except Exception:
            # 静默忽略，避免影响UI与既有流程
            pass


# ---------------------------------------------
# Excel 写入与绘图（由 GUI 负责，不在 core 中实现）
# ---------------------------------------------

def _write_led_sheets(
        wb: Workbook,
        led_index: int,
        result: Step1RawDataResult,
) -> None:
    """在汇总工作簿中为 LED{index} 写两个 Sheet。"""
    led_name = f"LED{led_index}"

    ws1 = wb.create_sheet(f"{led_name}_Category_Counts")
    ws1.append(["Value", "Count", "Proportion"])
    for e in result.category_stats:
        ws1.append([e.value, e.count, e.proportion])
    ws1.append([])
    ws1.append([f"读取到的数值单元格个数: {result.numbers_count}"])
    ws1.append([f"非数值（且非空）单元格个数: {result.non_numbers_count}"])
    ws1.append([f"空白单元格个数: {result.blank_cells}"])

    ws2 = wb.create_sheet(f"{led_name}_97Percent_Interval")
    ws2.append(["Min", "Median", "Max", "Interval_Prob"])
    ws2.append([result.min_val, result.median_val, result.max_val, result.interval_prob])


def _generate_line_chart(
        result: Step1RawDataResult,
        save_path: str,
        display_title: str,
) -> str | None:
    """根据分类统计生成折线图 PNG，返回保存路径；无数据返回 None。"""
    stats = result.category_stats
    if not stats:
        return None

    values = [e.value for e in stats]
    counts = [e.count for e in stats]
    if not values:
        return None

    # 归一化到 0..1 便于不同规模对比
    max_count = max(counts) if counts else 1
    norm_counts = [c / max_count for c in counts]

    plt.figure(figsize=(8, 5))
    plt.plot(values, norm_counts, linestyle="-", linewidth=1.2)
    plt.xlabel("Value")
    plt.ylabel("归一化 Count")
    plt.title(display_title)

    # 标注 min / median / max（落到最邻近采样点）
    def _nearest_idx(x: float) -> int:
        return min(range(len(values)), key=lambda i: abs(values[i] - x))

    for v, label in [
        (result.min_val, "Min"),
        (result.median_val, "Median"),
        (result.max_val, "Max"),
    ]:
        if v is None:
            continue
        idx = _nearest_idx(v)
        plt.scatter([values[idx]], [norm_counts[idx]], s=70, zorder=6, marker="D", label=label)
        plt.annotate(
            f"{label}\n{values[idx]}",
            xy=(values[idx], norm_counts[idx]),
            xytext=(0, 12),
            textcoords="offset points",
            ha="center",
            fontsize=9,
        )

    plt.grid(linestyle="--", alpha=0.6)
    plt.legend(loc="upper right")
    plt.tight_layout()

    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=220)
    plt.close()
    return save_path


# ---------------------------------------------
# PyQt Widget 类
# ---------------------------------------------
class Step1RawDataWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent

        # 参数
        self.target_prob = 0.97
        self.parse_string_numbers = True

        # 文件路径变量
        self.file_paths = [""] * 5

        # 图像标签引用
        self.image_labels = []
        self.image_paths = [""] * 5

        # 进度条控制
        self.progress_timer = QTimer()
        self.progress_timer.timeout.connect(self._update_progress)
        self.progress_value = 0

        self.setup_ui()

        # 不在这里初始化CSV，改为在showEvent中初始化
        self.csv_initialized = False

    def showEvent(self, event):
        """当界面显示时初始化CSV文件"""
        super().showEvent(event)
        if not self.csv_initialized:
            self.initialize_csv_files()
            self.csv_initialized = True

    def initialize_csv_files(self):
        """初始化CSV文件"""
        try:
            out_dir = _ensure_output_dir()
            print(f"初始化CSV文件，输出目录: {out_dir}")
            print(f"目录是否存在: {os.path.exists(out_dir)}")
            _init_led_csv_files(out_dir)
            print("CSV文件初始化完成")
        except Exception as e:
            print(f"初始化CSV文件失败: {e}")
            # 尝试再次创建目录
            try:
                out_dir = _ensure_output_dir()
                _init_led_csv_files(out_dir)
            except Exception as e2:
                print(f"重试初始化CSV文件失败: {e2}")


    def setup_ui(self):
        """设置UI界面"""
        # 主布局
        main_layout = QVBoxLayout(self)

        # 标题与提示
        title_label = QLabel("第一步：原始数据分析")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        description_text = (
            "为每个 LED 选择一个 Excel 文件（可留空）。\n"
            "程序会读取\n- 全部工作表\n- 全部单元格\n中的**数字**，并计算分类统计与覆盖≥97%的最短区间。\n"
            "结果将写入 data_files/step1_rawdata_analysis/step1_rawdata_analysis.xlsx，并生成每个 LED 的概率分布图。"
        )
        description_label = QLabel(description_text)
        description_label.setStyleSheet("color: #1f4a7c;")
        description_label.setWordWrap(True)
        main_layout.addWidget(description_label)

        # 选择文件区
        file_frame = QFrame()
        file_frame.setFrameStyle(QFrame.Box)
        file_layout = QVBoxLayout(file_frame)

        self.file_edits = []
        for i in range(5):
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)

            label = QLabel(f"LED{i + 1} 分bin电压数据路径：")
            label.setFixedWidth(200)
            row_layout.addWidget(label)

            file_edit = QLineEdit()
            file_edit.setReadOnly(True)
            row_layout.addWidget(file_edit)
            self.file_edits.append(file_edit)

            browse_btn = QPushButton("浏览")
            browse_btn.clicked.connect(lambda checked, idx=i: self._choose_file(idx))
            row_layout.addWidget(browse_btn)

            file_layout.addWidget(row_widget)

        main_layout.addWidget(file_frame)

        # 操作按钮
        button_layout = QHBoxLayout()
        run_button = QPushButton("写入数据")
        run_button.setStyleSheet("background-color: #d9ecff;")
        run_button.clicked.connect(self.on_run)
        button_layout.addWidget(run_button)
        button_layout.addStretch()
        main_layout.addLayout(button_layout)

        # 进度条（初始隐藏）
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setRange(0, 0)  # 不确定模式
        main_layout.addWidget(self.progress_bar)

        # 预览区：一行5个图像标签
        preview_layout = QHBoxLayout()
        self.image_labels = []
        for i in range(5):
            image_container = QFrame()
            image_container.setFrameStyle(QFrame.Box)
            image_container.setFixedSize(220, 160)
            image_layout = QVBoxLayout(image_container)

            label = QLabel(f"LED{i + 1}\n无图像")
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("color: #888; border: none;")
            label.mouseDoubleClickEvent = lambda event, idx=i: self._on_image_dblclick(idx)
            image_layout.addWidget(label)

            preview_layout.addWidget(image_container)
            self.image_labels.append(label)

        main_layout.addLayout(preview_layout)

        # 输出框
        output_label = QLabel("输出结果：")
        output_label.setAlignment(Qt.AlignLeft)
        main_layout.addWidget(output_label)

        self.output_box = QTextEdit()
        self.output_box.setReadOnly(True)
        main_layout.addWidget(self.output_box)

    def _choose_file(self, idx: int):
        """选择文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "请选择 Excel 文件,仅支持xlsx格式",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if file_path:
            self.file_paths[idx] = file_path
            self.file_edits[idx].setText(file_path)

    def _on_image_dblclick(self, idx: int):
        """双击图像标签"""
        out_dir = _ensure_output_dir()
        img_path = os.path.join(out_dir, f"LED{idx + 1}概率分布区间图.png")
        if not os.path.isfile(img_path):
            QMessageBox.warning(self, "提示", f"未找到文件：{img_path}")
            return
        self._open_file_in_os(img_path)

    def _open_file_in_os(self, path: str):
        """在操作系统中打开文件"""
        try:
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法打开文件：{e}")

    def _draw_image_on_label(self, idx: int, img_path: str | None):
        """在QLabel上显示图像"""
        label = self.image_labels[idx]
        if not img_path or not os.path.isfile(img_path):
            label.setText(f"LED{idx + 1}\n无图像")
            label.setStyleSheet("color: #888; border: none;")
            return

        try:
            pixmap = QPixmap(img_path)
            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(200, 140, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                label.setPixmap(scaled_pixmap)
                label.setStyleSheet("border: none;")
            else:
                label.setText(f"LED{idx + 1}\n图像加载失败")
                label.setStyleSheet("color: #c00; border: none;")
        except Exception:
            label.setText(f"LED{idx + 1}\n图像加载失败")
            label.setStyleSheet("color: #c00; border: none;")

    def _process_one(self, file_path: str, led_index: int, wb: Workbook, out_dir: str) -> str:
        """处理一个 LED：调用核心、写入Sheet、保存图像、绘制预览，返回文本摘要。"""
        result: Step1RawDataResult = process_raw_excel(
            file_path=file_path,
            target_prob=self.target_prob,
            parse_string_numbers=self.parse_string_numbers,
        )
        if not result.has_data:
            # 仍写入空统计，以便汇总结构完整
            _write_led_sheets(wb, led_index, result)
            self._draw_image_on_label(led_index - 1, None)
            return f"[LED{led_index}] 未在工作簿中找到可用数字，已写入空统计。\n"

        # 写两个Sheet
        _write_led_sheets(wb, led_index, result)

        # 保存图像并绘制到Label
        chart_path = os.path.join(out_dir, f"LED{led_index}概率分布区间图.png")
        _generate_line_chart(result, chart_path, f"LED{led_index} 概率分布图")
        self._draw_image_on_label(led_index - 1, chart_path)

        # 将该 Excel 的纯数字写入对应CSV（使用 step1_rawdata_analysis_2.read_excel_numeric）
        try:
            data_numeric = read_excel_numeric(file_path)
            # 扁平化为一列
            flat_numbers = [float(x) for row in data_numeric for x in row if x is not None]
            if flat_numbers:
                raw_csv_path = os.path.join(out_dir, f"LED{led_index}_raw_data.csv")
                with open(raw_csv_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    for num in flat_numbers:
                        writer.writerow([num])
        except Exception:
            # 静默忽略写入错误
            pass

        # 文本摘要
        summary = (
            f"[LED{led_index}] 完成：97%区间 Min={result.min_val}, Median={result.median_val}, Max={result.max_val}；"
            f"覆盖概率≈{result.interval_prob:.4f}\n  - 分布图：{chart_path}\n"
        )
        return summary

    def _update_progress(self):
        """更新进度条动画"""
        self.progress_value += 1
        if self.progress_value >= 100:
            self.progress_value = 0

    def on_run(self):
        """执行分析"""
        paths = self.file_paths.copy()
        out_dir = _ensure_output_dir()
        out_xlsx = os.path.join(out_dir, "step1_rawdata_analysis.xlsx")

        # 清空输出框 & 图像标签
        self.output_box.clear()
        for i in range(5):
            self._draw_image_on_label(i, None)

        # 覆盖旧文件
        if os.path.exists(out_xlsx):
            try:
                os.remove(out_xlsx)
            except Exception:
                pass

        wb = Workbook()
        # 移除默认 Sheet
        try:
            default_sheet = wb.active
            wb.remove(default_sheet)
        except Exception:
            pass

        any_ok = False

        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_timer.start(50)  # 20fps动画

        # 处理每个LED文件
        for i, p in enumerate(paths, start=1):
            if not p:
                self.output_box.append(f"[LED{i}] 未选择文件，跳过。")
                continue
            if not os.path.isfile(p):
                self.output_box.append(f"[LED{i}] 文件不存在：{p}")
                continue
            try:
                summary = self._process_one(p, i, wb, out_dir)
                self.output_box.append(summary)
                any_ok = True
            except Exception as e:
                self.output_box.append(f"[LED{i}] 处理失败：{e}")

        # 停止进度条动画
        self.progress_timer.stop()
        self.progress_bar.setVisible(False)

        if any_ok:
            # 保存汇总 Excel
            try:
                wb.save(out_xlsx)
                self.output_box.append(f"\n汇总Excel已输出：{out_xlsx}")
                QMessageBox.information(self, "完成", "原始数据分析已完成。")
            except Exception as e:
                msg = f"保存汇总Excel失败：{e}"
                self.output_box.append(f"\n【导出失败】{msg}")
                QMessageBox.critical(self, "错误", msg)
                return
        else:
            self.output_box.append("\n没有任何有效文件被处理，请检查Excel内容是否包含数字。")
            QMessageBox.warning(self, "提示", "没有任何有效文件被处理，请检查Excel内容是否包含数字。")


# ---------------------------------------------
# 独立运行入口
# ---------------------------------------------

class Step1RawDataWindow(QWidget):
    """独立运行时的窗口类，保持向后兼容"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("第一步：原始数据分析（读取全表全单元格）")
        self.resize(1280, 900)

        # 创建主Widget
        layout = QVBoxLayout(self)
        self.main_widget = Step1RawDataWidget(self)
        layout.addWidget(self.main_widget)


def main():
    app = QApplication(sys.argv)
    window = Step1RawDataWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()